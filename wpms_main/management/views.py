import math
import os
from datetime import timedelta

import openpyxl
from docxtpl import DocxTemplate
from django.contrib.auth import authenticate, logout, login
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404

# Create your views here.
from django.http import HttpResponse
from django.template import loader
from django.utils import timezone
from .models import *
from .forms import *

def index(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/index.html")
    context = {"user": Users.objects.filter(username=request.user.username).get()}
    return HttpResponse(template.render(context, request))

def prod(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/prod.html")
    table_data = []
    for nom in ProductNom.objects.all():
        coming = ProductSpecification.objects.filter(product_nom_id=nom.id)
        outcoming = PenalSpecification.objects.filter(status__in=["done", "archive"]).filter(product_nom_id=nom.id)
        if not coming:
            continue
        weight = 0
        for c in coming:
            weight += c.weight
        for o in outcoming:
            weight -= sum(o.weight_list)
        if weight <= 0:
            continue
        spec = {}
        spec["nums"] = " ,".join(set(map(str, map(lambda x: x.id, coming))))
        spec["sender"] = " ,".join(set(map(str, map(lambda x: x.sender, coming))))
        spec["nom"] = str(nom)
        spec["weight"] = weight
        table_data.append(spec)
    total = sum(map(lambda x: int(x["weight"]), table_data))
    context = {
        "user": Users.objects.filter(username=request.user.username).get(),
        "table_data": table_data,
        "total": total
    }
    return HttpResponse(template.render(context, request))

def production(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/production.html")
    context = {"user": Users.objects.filter(username=request.user.username).get()}
    return HttpResponse(template.render(context, request))

def productiondocuments(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/proddocs.html")
    doc_data = []
    for spec in ProductSpecification.objects.all():
        data = {}
        data["id"] = spec.id
        data["type"] = "Спецификация продукции"
        spec_date = spec.date + timedelta(hours=3)
        data["date"] = spec_date.strftime("%d.%m.%Y %H:%M")
        data["sender"] = " ".join(str(spec.sender).split()[1:])
        data["weight"] = spec.weight
        data["link"] = f"/productspecification/{spec.id}/generate/prodspec{spec.id}.xlsx"
        data['action_text'] = "Сгенерировать xlsx"
        doc_data.append(data)
    context = {
        "user": Users.objects.filter(username=request.user.username).get(),
        "doc_data": doc_data
    }
    return HttpResponse(template.render(context, request))

def generate_product_specification(request, id, fn):
    if not request.user.is_authenticated:
        return redirect("/login")
    tmplt_path = os.path.join(os.path.dirname(__file__), 'static', 'files', 'product_specification_template.xlsx')
    res_path = os.path.join(os.path.dirname(__file__), 'static', 'files', 'product_specification.xlsx')
    spec = get_object_or_404(ProductSpecification, pk=id)
    data = {}
    data["date"] = f"от {spec.date.strftime('%d.%m.%Y')} г. ДЕНЬ/НОЧЬ"
    data["brigade_num"] = f"смена {spec.brigade_num}"
    data["weight_list"] = []
    data["total"] = spec.weight
    for i in range(1, spec.kip_count):
        weight_data = {}
        weight_data["num"] = i
        weight_data["color"] = str(spec.product_nom)
        weight_data["weight"] = round(spec.weight / spec.kip_count)
        data["weight_list"].append(weight_data)
    sum_weight = round(spec.weight / spec.kip_count) * (spec.kip_count - 1)
    last_weight = spec.weight - sum_weight
    data["weight_list"].append({
        "num": spec.kip_count,
        "color": str(spec.product_nom),
        "weight": last_weight
    })
    generate_product_specification_xls(tmplt_path, res_path, data)
    with open(res_path, "rb") as file:
        response = HttpResponse(content=file)
        response['Content-Type'] = 'application/xlsx'
    return response

def generate_product_specification_xls(tmplt_path, res_path, data):
    wb = openpyxl.load_workbook(tmplt_path)
    pg = wb['Лист3']
    pg["B4"] = data["date"]
    pg["G8"] = data["brigade_num"]
    for w in data["weight_list"]:
        pg.append(["", w["num"], w["color"], w["weight"]])
    pg.append(["", "", "ИТОГО", data["total"]])
    pg.append([""])
    pg.append(["", "", "Должность", "ФИО", "Подпись"])
    pg.append([""])
    pg.append([""])
    pg.append([""])
    pg.append([""])
    wb.save(res_path)

def store(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    doc_data = []
    template = loader.get_template("management/documents.html")
    for act in AcceptanceAct.objects.filter(~Q(status="archive")):
        doc = {}
        doc["id"] = act.act_num
        doc["type"] = "Акт о приеме сырья"
        act_date = act.date + timedelta(hours=3)
        doc["date"] = act_date.strftime("%d.%m.%Y %H:%M")
        doc["sender"] = " ".join(str(act.sender).split()[1:])
        doc["receiver"] = " ".join(str(act.receiver).split()[1:])
        doc["obj_id"] = act.id
        doc["archive_link"] = f"acceptanceact/{act.id}/archive"
        doc["delete_link"] = f"acceptanceact/{act.id}/delete"
        if act.status == "new":
            if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (
                    Users.objects.filter(username=request.user.username).get().role == "worker"):
                continue
            doc["link"] = f"/acceptanceact/{act.id}/init"
            doc["action_text"] = "Добавить данные акта"
            doc["status_class"] = "not-done"
            doc["status_text"] = "Ввод данных"
        if act.status == "init":
            if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (
                    Users.objects.filter(username=request.user.username).get().role == "worker"):
                continue
            doc["link"] = f"/acceptanceact/{act.id}/weight"
            doc["action_text"] = "Указать количество кип"
            doc["status_class"] = "not-done"
            doc["status_text"] = "Подсчет кип"
        if act.status == "weight":
            if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (
                    Users.objects.filter(username=request.user.username).get().role == "security"):
                continue
            doc["link"] = f"/acceptanceact/{act.id}/carweight"
            doc["action_text"] = "Добавить вес автомобиля"
            doc["status_class"] = "not-done"
            doc["status_text"] = "Взвешивание автомобиля"
        if act.status == "secondmaterial":
            if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (
                    Users.objects.filter(username=request.user.username).get().role == "worker"):
                continue
            doc["link"] = f"/acceptanceact/{act.id}/secondmaterial"
            doc["action_text"] = "Добавить второй вид сырья"
            doc["link2"] = f"/acceptanceact/{act.id}/skip"
            doc["action2_text"] = "Пропустить"
            doc["status_class"] = "not-done"
            doc["status_text"] = "Подсчет кип"
        if act.status == "done":
            if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (
                    Users.objects.filter(username=request.user.username).get().role == "worker"):
                continue
            doc["link"] = f"/acceptanceact/{act.id}/generate/act{act.act_num}.xls"
            doc["action_text"] = "Сгенерировать xls"
            doc["status_class"] = "done"
            doc["status_text"] = "Оформление завершено"
        doc_data.append(doc)
    for sp in PenalSpecification.objects.filter(~Q(status="archive")):
        doc = {}
        doc["id"] = sp.specification_num
        doc["type"] = "Спецификация пенала"
        sp_date = sp.date + timedelta(hours=3)
        doc["date"] = sp_date.strftime("%d.%m.%Y %H:%M")
        doc["sender"] = " ".join(str(sp.sender).split()[1:])
        doc["receiver"] = " ".join(str(sp.receiver).split()[1:])
        doc["obj_id"] = sp.id
        doc["archive_link"] = f"penalspecification/{sp.id}/archive"
        doc["delete_link"] = f"penalspecification/{sp.id}/delete"
        if sp.status == "new":
            if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (
                    Users.objects.filter(username=request.user.username).get().role == "worker"):
                continue
            doc["link"] = f"/penalspecification/{sp.id}/init"
            doc["action_text"] = "Указать данные спецификации"
            doc["status_class"] = "not-done"
            doc["status_text"] = "Ввод данных"
        if sp.status == "init":
            if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (
                    Users.objects.filter(username=request.user.username).get().role == "worker"):
                continue
            doc["link"] = f"/penalspecification/{sp.id}/weight"
            doc["action_text"] = "Указать веса б/б"
            doc["link2"] = f"/penalspecification/{sp.id}/count"
            doc["action2_text"] = "Указать количество б/б"
            doc["status_class"] = "not-done"
            doc["status_text"] = "Взвешивание пенала"
        if sp.status == "carweight":
            if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (
                    Users.objects.filter(username=request.user.username).get().role == "security"):
                continue
            doc["link"] = f"/penalspecification/{sp.id}/carweight"
            doc["action_text"] = "Добавить вес автомобиля"
            doc["status_class"] = "not-done"
            doc["status_text"] = "Взвешвание автомобиля"
        if sp.status == "done":
            if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (
                    Users.objects.filter(username=request.user.username).get().role == "worker"):
                continue
            doc["link"] = f"/penalspecification/{sp.id}/generate/spec{sp.specification_num}.xls"
            doc["action_text"] = "Сгенерировать xls"
            doc["status_class"] = "done"
            doc["status_text"] = "Оформление завершено"
        doc_data.append(doc)
    for wb in Waybill.objects.filter(~Q(status="archive")):
        if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (
                Users.objects.filter(username=request.user.username).get().role == "worker"):
            continue
        doc = {}
        doc["id"] = wb.waybill_num
        doc["obj_id"] = wb.id
        doc["type"] = "Транспортная накладная"
        wb_date = wb.date + timedelta(hours=3)
        doc["date"] = wb_date.strftime("%d.%m.%Y %H:%M")
        doc["sender"] = " ".join(str(wb.specification.sender).split()[1:])
        doc["receiver"] = " ".join(str(wb.specification.receiver).split()[1:])
        doc["archive_link"] = f"waybill/{wb.id}/archive"
        doc["delete_link"] = f"waybill/{wb.id}/delete"
        doc["link"] = f"/waybill/{wb.id}/generate/waybill{wb.waybill_num}.xls"
        doc["action_text"] = "Сгенерировать xls"
        doc["status_class"] = "done"
        doc["status_text"] = "Оформление завершено"
        doc_data.append(doc)
    doc_data.sort(key=lambda x: x["date"], reverse=True)
    context = {
        "doc_data": doc_data,
        "user": Users.objects.filter(username=request.user.username).get(),
        "transfer": TransferToProd.objects.all(),
        "transfer_link": f"/transfer/transfer_to_prod{timezone.now().day}.{timezone.now().month}.{timezone.now().year}.xlsx"
    }
    return HttpResponse(template.render(context, request))

def handbook(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/handbooks.html")
    model_data = [{
            "name": "Организации",
            "count": Organization.objects.all().count(),
            "link": "/organization",
            "entities": list(map(lambda x: {
                "href": "/organization/" + str(x.id),
                "value": str(x)
            }, Organization.objects.all()))
        },
        {
            "name": "Виды сырья",
            "count": RawMaterial.objects.all().count(),
            "link": "/rawmaterial",
            "entities": list(map(lambda x: {
                "href": "/rawmaterial/" + str(x.id),
                "value": str(x)
            }, RawMaterial.objects.all()))
        },
        {
            "name": "Автомобили",
            "count": Car.objects.all().count(),
            "link": "/car",
            "entities": list(map(lambda x: {
                "href": "/car/" + str(x.id),
                "value": str(x)
            }, Car.objects.all()))
        },
        {
            "name": "Водители",
            "count": Driver.objects.all().count(),
            "link": "/driver",
            "entities": list(map(lambda x: {
                "href": "/driver/" + str(x.id),
                "value": str(x)
            }, Driver.objects.all()))
        },
        {
            "name": "Сотрудники",
            "count": Worker.objects.all().count(),
            "link": "/worker",
            "entities": list(map(lambda x: {
                "href": "/worker/" + str(x.id),
                "value": str(x)
            }, Worker.objects.all()))
        },
        {
            "name": "Характеристики продукции",
            "count": ProductFeature.objects.all().count(),
            "link": "/productfeature",
            "entities": list(map(lambda x: {
                "href": "/productfeature/" + str(x.id),
                "value": str(x)
            }, ProductFeature.objects.all()))
        },
        {
            "name": "Номенклатуры продукции",
            "count": ProductNom.objects.all().count(),
            "link": "/productnom",
            "entities": list(map(lambda x: {
                "href": "/productnom/" + str(x.id),
                "value": str(x)
            }, ProductNom.objects.all()))
        },
    ]
    context = {
        "model_data": model_data
    }
    return HttpResponse(template.render(context, request))

def archive(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    doc_data = []
    template = loader.get_template("management/archive.html")
    for act in AcceptanceAct.objects.filter(status="archive"):
        doc = {}
        doc["id"] = act.act_num
        doc["type"] = "Акт о приеме сырья"
        doc_date = act.date + timedelta(hours=3)
        doc["date"] = doc_date.strftime("%d.%m.%Y %H:%M")
        doc["sender"] = " ".join(str(act.sender).split()[1:])
        doc["receiver"] = " ".join(str(act.receiver).split()[1:])
        doc["obj_id"] = act.id
        doc["archive_link"] = f"acceptanceact/{act.id}/archive"
        doc["delete_link"] = f"acceptanceact/{act.id}/delete"
        doc["link"] = f"/acceptanceact/{act.id}/generate/act{act.act_num}.xls"
        doc["action_text"] = "Сгенерировать xls"
        doc["status_class"] = "done"
        doc["status_text"] = "Оформление завершено"
        doc_data.append(doc)
    for sp in PenalSpecification.objects.filter(status="archive"):
        doc = {}
        doc["id"] = sp.specification_num
        doc["type"] = "Спецификация пенала"
        sp_date = sp.date + timedelta(hours=3)
        doc["date"] = sp_date.strftime("%d.%m.%Y %H:%M")
        doc["sender"] = " ".join(str(sp.sender).split()[1:])
        doc["receiver"] = " ".join(str(sp.receiver).split()[1:])
        doc["obj_id"] = sp.id
        doc["archive_link"] = f"penalspecification/{sp.id}/archive"
        doc["delete_link"] = f"penalspecification/{sp.id}/delete"
        doc["link"] = f"/penalspecification/{sp.id}/generate/spec{sp.specification_num}.xls"
        doc["action_text"] = "Сгенерировать xls"
        doc["status_class"] = "done"
        doc["status_text"] = "Оформление завершено"
        doc_data.append(doc)
    for wb in Waybill.objects.filter(status="archive"):
        doc = {}
        doc["id"] = wb.waybill_num
        doc["obj_id"] = wb.id
        doc["type"] = "Транспортная накладная"
        wb_date = wb.date + timedelta(hours=3)
        doc["date"] = wb_date.strftime("%d.%m.%Y %H:%M")
        doc["sender"] = " ".join(str(wb.specification.sender).split()[1:]) if wb.specification else "-"
        doc["receiver"] = " ".join(str(wb.specification.receiver).split()[1:]) if wb.specification else "-"
        doc["archive_link"] = f"waybill/{wb.id}/archive"
        doc["delete_link"] = f"waybill/{wb.id}/delete"
        doc["link"] = f"/waybill/{wb.id}/generate/waybill{wb.waybill_num}.xls"
        doc["action_text"] = "Сгенерировать xls"
        doc["status_class"] = "done"
        doc["status_text"] = "Оформление завершено"
        doc_data.append(doc)
    doc_data.sort(key=lambda x: x["date"], reverse=True)
    context = {
        "doc_data": doc_data,
        "user": Users.objects.filter(username=request.user.username).get()
    }
    return HttpResponse(template.render(context, request))

def acceptance_act_archive(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    if request.method == "GET":
        act = get_object_or_404(AcceptanceAct, pk=id)
        act.status = "archive"
        act.save()
        return HttpResponse('success')


def penal_specification_archive(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    if request.method == "GET":
        act = get_object_or_404(PenalSpecification, pk=id)
        act.status = "archive"
        act.save()
        return HttpResponse('success')

def waybill_archive(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    if request.method == "GET":
        act = get_object_or_404(Waybill, pk=id)
        act.status = "archive"
        act.save()
        return HttpResponse('success')

def acceptance_act_delete(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    if not Users.objects.filter(username=request.user.username).get().role in "superadmin":
        return redirect("/")
    if request.method == "GET":
        act = get_object_or_404(AcceptanceAct, pk=id)
        act.delete()
        return HttpResponse('success')


def penal_specification_delete(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    if not Users.objects.filter(username=request.user.username).get().role in "superadmin":
        return redirect("/")
    if request.method == "GET":
        act = get_object_or_404(PenalSpecification, pk=id)
        act.delete()
        return HttpResponse('success')

def waybill_delete(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    if not Users.objects.filter(username=request.user.username).get().role in "superadmin":
        return redirect("/")
    if request.method == "GET":
        act = get_object_or_404(Waybill, pk=id)
        act.delete()
        return HttpResponse('success')

def waybill(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/form.html")
    header = "Добавить транспортную накладную"
    if request.method == "POST":
        form = WaybillForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("/store")
        else:
            header = "Ошибка при сохранении"
    form = WaybillForm()
    context = {
        "form": form,
        "url": "waybill",
        "header": header,
        "href": "/store",
        "is_doc": True
    }
    return HttpResponse(template.render(context, request))

def generate_transfer_xls(tmplt_path, res_path, penal_data, transfer_data):
    print(penal_data)
    print(transfer_data)
    wb = openpyxl.load_workbook(tmplt_path)
    pg = wb['Лист1']
    for penal in penal_data:
        pg.append([penal[0], "", "", "", penal[1], penal[2], penal[3]])
    pg.append(["№ смены", "Дата", "время", "№ пенала", "Вес выданной бутылки", "Кол-во кип", "ФИО ответственных"])
    for transfer in transfer_data:
        pg.append([transfer[0], transfer[1], transfer[2], transfer[3], transfer[4], transfer[5]])
    wb.save(res_path)

def generate_transfer(request, fn):
    if not request.user.is_authenticated:
        return redirect("/login")
    tmplt_path = os.path.join(os.path.dirname(__file__), 'static', 'files', 'transfer_template.xlsx')
    res_path = os.path.join(os.path.dirname(__file__), 'static', 'files', 'transfer.xlsx')
    penal_data = []
    transfer_data = []
    act_nums = set()
    for transfer in TransferToProd.objects.all():
        p = []
        act = AcceptanceAct.objects.get(pk=transfer.act_id)
        if act.act_num not in act_nums and act.weight_list:
            kip_count = act.kip_count if not act.kip_count2 else act.kip_count + act.kip_count2
            p.append(act.act_num)
            p.append(sum(act.weight_list))
            p.append(kip_count)
            p.append(sum(act.weight_list) / kip_count)
            penal_data.append(p)
            act_nums.add(act.act_num)
        t = []
        t.append(transfer.shift_num)
        transfer_date = transfer.date + timedelta(hours=3)
        t.append(transfer_date.strftime("%d.%m.%Y"))
        t.append(transfer_date.strftime("%H:%M"))
        t.append(act.act_num)
        t.append(transfer.transfer_weight)
        t.append(transfer.transfer_count)
        transfer_data.append(t)
    generate_transfer_xls(tmplt_path, res_path, penal_data, transfer_data)
    with open(res_path, "rb") as file:
        response = HttpResponse(content=file)
        response['Content-Type'] = 'application/xlsx'
    return response

def generate_waybill_xls(tmplt_path, res_path, data):
    wb = openpyxl.load_workbook(tmplt_path)
    pg1 = wb['стр.1']
    pg2 = wb["стр.2"]
    pg1["G7"] = data["date"]
    pg1["AC7"] = data["num"]
    pg1["B13"] = data["sender"]
    pg1["BP13"] = data["receiver"]
    pg1["B23"] = data["feature"].lower()
    pg1["BF23"] = data["count"]
    pg1["B25"] = data["weight"]
    pg1["BF42"] = data["driver"].upper()
    pg1["B45"] = data["mark"].upper()
    pg1["BF45"] = data["reg_num"].upper()
    pg1["B58"] = data["sender"]
    pg1["BF58"] = data["date"]
    pg1["B60"] = data["date"]
    pg1["BF60"] = data["date"]
    pg1["BE66"] = data["weight"]
    pg1["BR68"] = data["driver"].upper()
    pg2["BR11"] = data["driver"].upper()
    wb.save(res_path)

def generate_waybill(request, id, fn):
    if not request.user.is_authenticated:
        return redirect("/login")
    tmplt_path = os.path.join(os.path.dirname(__file__), 'static', 'files', 'waybill_template.xlsx')
    res_path = os.path.join(os.path.dirname(__file__), 'static', 'files', 'waybill.xlsx')
    wb = get_object_or_404(Waybill, pk=id)
    data = {}
    wb_date = wb.date + timedelta(hours=3)
    data["date"] = wb_date.strftime("%d.%m.%Y")
    data["num"] = wb.waybill_num
    data["sender"] = f'{wb.specification.sender.type}"{wb.specification.sender.name}" {wb.specification.sender.address}'
    data["receiver"] = f'{wb.specification.receiver.type}"{wb.specification.receiver.name}" {wb.specification.receiver.address}'
    data["feature"] = str(wb.specification.product_nom)
    data["count"] = len(wb.specification.weight_list)
    data["weight"] = wb.specification.out_weight - wb.specification.in_weight
    data["driver"] = str(wb.driver)
    data["mark"] = str(wb.car.mark)
    data["reg_num"] = str(wb.car.reg_num)
    generate_waybill_xls(tmplt_path, res_path, data)
    with open(res_path, "rb") as file:
        response = HttpResponse(content=file)
        response['Content-Type'] = 'application/xlsx'
    return response


def penal_specification(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/form.html")
    header = "Добавить спецификацию пенала"
    if request.method == "POST":
        form = PenalSpecificationForm(request.POST)
        if form.is_valid():
            act = form.save(commit=False)
            act.status = "new"
            act.save()
            return redirect("/store")
        else:
            header = "Ошибка при сохранении"
    form = PenalSpecificationForm()
    context = {
        "form": form,
        "url": "penalspecification",
        "header": header,
        "href": "/store",
        "is_doc": True
    }
    return HttpResponse(template.render(context, request))

def penal_specification_init(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    spec = get_object_or_404(PenalSpecification, pk=id)
    template = loader.get_template("management/form.html")
    header = "Добавить данные спецификации"
    if request.method == "POST":
        form = PenalSpecificationFormInit(request.POST, instance=spec)
        if form.is_valid():
            act = form.save(commit=False)
            act.status = "init"
            act.save()
            return redirect("/store")
        else:
            header = "Ошибка при сохранении"
    if Organization.objects.filter(is_default=True).first():
        form = PenalSpecificationFormInit(initial={'sender': Organization.objects.filter(is_default=True).first()})
    else:
        form = PenalSpecificationFormInit()
    context = {
        "form": form,
        "url": f"/penalspecification/{id}/init",
        "header": header,
        "href": "/store",
        "is_doc": True
    }
    return HttpResponse(template.render(context, request))

def penal_specification_weight(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    spec = get_object_or_404(PenalSpecification, pk=id)
    template = loader.get_template("management/form.html")
    header = "Добавить веса б/б"
    if request.method == "POST":
        form = PenalSpecificationWeightsForm(request.POST, instance=spec)
        if form.is_valid():
            spec = form.save(commit=False)
            spec.status = "carweight"
            spec.save()
            return redirect("/store")
        else:
            header = "Убедитесь в корректном заполнении списка весов"
    form = PenalSpecificationWeightsForm()
    context = {
        "form": form,
        "url": f"/penalspecification/{id}/weight",
        "header": header,
        "href": "/store"
    }
    return HttpResponse(template.render(context, request))

def penal_specification_count(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    spec = get_object_or_404(PenalSpecification, pk=id)
    template = loader.get_template("management/form.html")
    header = "Добавить количество упаковок"
    if request.method == "POST":
        form = PenalSpecificationCountForm(request.POST, instance=spec)
        if form.is_valid():
            spec = form.save(commit=False)
            spec.status = "carweight"
            spec.save()
            return redirect("/store")
        else:
            header = "Ошибка при вводе данных"
    form = PenalSpecificationCountForm()
    context = {
        "form": form,
        "url": f"/penalspecification/{id}/count",
        "header": header,
        "href": "/store"
    }
    return HttpResponse(template.render(context, request))

def penal_specification_carweight(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    spec = get_object_or_404(PenalSpecification, pk=id)
    template = loader.get_template("management/form.html")
    header = "Добавить вес автомобиля на выезде"
    if request.method == "POST":
        form = PenalSpecificationCarWeightForm(request.POST, instance=spec)
        if form.is_valid():
            spec = form.save(commit=False)
            spec.status = "done"
            valid = True
            if not spec.weight_list:
                weight = []
                bag_weight = round((int(spec.out_weight) - int(spec.in_weight)) / spec.kip_count)
                bag_count = spec.kip_count
                while bag_count > 0:
                    weight.append(bag_weight)
                    bag_count -= 1
                spec.weight_list = weight
            else:
                real_weight = int(spec.out_weight) - int(spec.in_weight)
                if sum(spec.weight_list) > real_weight:
                    valid = False
                delta = real_weight - sum(spec.weight_list)
                weight = []
                if delta > 1:
                    added_weight = delta / len(spec.weight_list)
                    for w in spec.weight_list:
                        weight.append(round(added_weight + w))
                    spec.weight_list = weight
            if valid:
                spec.save()
                return redirect("/store")
            else:
                header = "Ошибка. Веса б/б больше чем разница весов автомобиля."
        else:
            header = "Ошибка при вводе данных"
    form = PenalSpecificationCarWeightForm()
    context = {
        "form": form,
        "url": f"/penalspecification/{id}/carweight",
        "header": header,
        "href": "/store"
    }
    return HttpResponse(template.render(context, request))

def generate_penal_specification_xls(tmplt_path, res_path, data):
    wb = openpyxl.load_workbook(tmplt_path)
    pg = wb['Лист1']
    pg["D3"] = data["num"]
    pg["E3"] = data["date"]
    pg["D5"] = data["sender"].upper()
    pg["D6"] = data["receiver"].upper()
    i = 1
    for w in data["weights"]:
        pg.append([i, data["feature"], " ", data["nom"], " ", w])
        i += 1
    pg.append([" ", " ", " ", " ", " "])
    pg.append([i - 1, " ", " ", "ВСЕГО", " ", sum(data["weights"])])
    pg.append(["Отпустил кладовщик:", " ", " ", " ", data["sending_worker"].upper()])
    pg.append([" "])
    pg.append(["Получил"])
    wb.save(res_path)

def generate_penal_specification(request, id, fn):
    if not request.user.is_authenticated:
        return redirect("/login")
    tmplt_path = os.path.join(os.path.dirname(__file__), 'static', 'files', 'penal_specification_template.xlsx')
    res_path = os.path.join(os.path.dirname(__file__), 'static', 'files', 'penal_specification.xlsx')
    sp = get_object_or_404(PenalSpecification, pk=id)
    data = {}
    data["num"] = sp.specification_num
    sp_date = sp.date + timedelta(hours=3)
    data["date"] = sp_date.strftime("%d.%m.%Y")
    data["sender"] = f'{sp.sender.type}"{sp.sender.name}"'
    data["receiver"] = f'{sp.receiver.type}"{sp.receiver.name}"'
    data["weights"] = sp.weight_list
    data["feature"] = str(sp.product_feature)
    data["nom"] = str(sp.product_nom)
    data["sending_worker"] = str(sp.sending_worker)
    generate_penal_specification_xls(tmplt_path, res_path, data)
    with open(res_path, "rb") as file:
        response = HttpResponse(content=file)
        response['Content-Type'] = 'application/xlsx'
    return response

def acceptanceact(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (Users.objects.filter(username=request.user.username).get().role == "security"):
        return redirect("/")
    template = loader.get_template("management/form.html")
    header = "Добавить акт приема сырья"
    if request.method == "POST":
        form = AcceptanceActForm(request.POST)
        if form.is_valid():
            act = form.save(commit=False)
            act.status = "new"
            act.save()
            return redirect("/store")
        else:
            header = "Ошибка при сохранении"
    form = AcceptanceActForm()
    context = {
        "form": form,
        "url": "acceptanceact",
        "header": header,
        "href": "/store",
        "is_doc": True
    }
    return HttpResponse(template.render(context, request))

def acceptanceactinit(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (Users.objects.filter(username=request.user.username).get().role == "worker"):
        return redirect("/")
    template = loader.get_template("management/form.html")
    act = get_object_or_404(AcceptanceAct, pk=id)
    header = "Указать данные акта приема сырья"
    if request.method == "POST":
        form = AcceptanceActFormInit(request.POST, instance=act)
        if form.is_valid():
            act = form.save(commit=False)
            act.status = "init"
            act.save()
            return redirect("/store")
        else:
            header = "Ошибка при сохранении"
    if Organization.objects.filter(is_default=True).first():
        form = AcceptanceActFormInit(initial={'receiver': Organization.objects.filter(is_default=True).first()})
    else:
        form = AcceptanceActFormInit()
    context = {
        "form": form,
        "url": f"/acceptanceact/{id}/init",
        "header": header,
        "href": "/store",
        "is_doc": True
    }
    return HttpResponse(template.render(context, request))

def acceptanceactweight(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (Users.objects.filter(username=request.user.username).get().role == "worker"):
        return redirect("/")
    act = get_object_or_404(AcceptanceAct, pk=id)
    template = loader.get_template("management/form.html")
    header = "Добавить количество кип"
    if request.method == "POST":
        form = AcceptanceActWeightsForm(request.POST, instance=act)
        if form.is_valid():
            act = form.save(commit=False)
            act.status = "secondmaterial"
            act.penal_count = 3
            act.save()
            return redirect("/store")
        else:
            header = "Убедитесь в корректном заполнении формы"
    form = AcceptanceActWeightsForm()
    context = {
        "form": form,
        "url": f"/acceptanceact/{id}/weight",
        "header": header,
        "href": "/store",
        "hint": "Если в акте присутствует несколько видов сырья, заполняйте здесь только для того вида, который был указан основным"
    }
    return HttpResponse(template.render(context, request))

def acceptanceactsecondmaterial(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (Users.objects.filter(username=request.user.username).get().role == "worker"):
        return redirect("/")
    act = get_object_or_404(AcceptanceAct, pk=id)
    template = loader.get_template("management/form.html")
    header = "Добавить второй вид сырья"
    if request.method == "POST":
        form = AcceptanceActSecondMaterialForm(request.POST, instance=act)
        if form.is_valid():
            act = form.save(commit=False)
            act.status = "weight"
            act.penal_count2 = 3
            act.save()
            return redirect("/store")
        else:
            header = "Убедитесь в корректном заполнении списка весов"
    form = AcceptanceActSecondMaterialForm()
    context = {
        "form": form,
        "url": f"/acceptanceact/{id}/secondmaterial",
        "header": header,
        "href": "/store"
    }
    return HttpResponse(template.render(context, request))

def acceptance_act_skip_second_material(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (Users.objects.filter(username=request.user.username).get().role == "worker"):
        return redirect("/")
    if request.method == "GET":
        act = get_object_or_404(AcceptanceAct, pk=id)
        act.status = "weight"
        act.save()
        return redirect("/store")

def acceptanceactcarweight(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    if not Users.objects.filter(username=request.user.username).get().role in "superadmin" and not (Users.objects.filter(username=request.user.username).get().role == "security"):
        return redirect("/")
    act = get_object_or_404(AcceptanceAct, pk=id)
    template = loader.get_template("management/form.html")
    header = "Добавить вес автомобиля на выезде"
    if request.method == "POST":
        form = AcceptanceActCarWeightsForm(request.POST, instance=act)
        if form.is_valid():
            act = form.save(commit=False)
            act.status = "done"
            weight = []
            kip_weight = round((int(act.in_weight) - int(act.out_weight)) / act.kip_count) if not act.kip_count2 else round((int(act.in_weight) - int(act.out_weight)) / (act.kip_count + act.kip_count2))
            kip_count = act.kip_count
            kip_count2 = act.kip_count2
            i = act.penal_count
            while kip_count > 0:
                if kip_count // i > 0:
                    kip_count -= i
                    weight.append(i * kip_weight)
                else:
                    weight.append(kip_count * kip_weight)
                    kip_count -= i
            if kip_count2:
                i = act.penal_count2
                while kip_count2 > 0:
                    if kip_count2 // i > 0:
                        kip_count2 -= i
                        weight.append(i * kip_weight)
                    else:
                        weight.append(kip_count2 * kip_weight)
                        kip_count2 -= i
            act.weight_list = weight
            act.save()
            return redirect("/store")
        else:
            header = "Ошибка при сохранении"
    form = AcceptanceActCarWeightsForm()
    context = {
        "form": form,
        "url": f"/acceptanceact/{id}/carweight",
        "header": header,
        "href": "/store",
        "is_doc": True
    }
    return HttpResponse(template.render(context, request))

def transfer(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/form.html")
    header = "Отправить в производство"
    if request.method == "POST":
        form = TransferToProdForm(request.POST)
        valid = True
        if form.is_valid():
            act = AcceptanceAct.objects.get(pk=form.cleaned_data["act"].id)
            remain_count = act.kip_count if not act.kip_count2 else act.kip_count + act.kip_count2
            for transfer in TransferToProd.objects.filter(act_id=act.id):
                remain_count -= transfer.transfer_count
            if remain_count < form.cleaned_data["transfer_count"]:
                valid = False
                header = "Ошибка. Количество кип превышает оставшиеся в партии."
            if valid:
                form.save()
                return redirect("/store")
        else:
            header = "Ошибка при сохранении"
    bank_form = TransferToProdForm()
    context = {
        "form": bank_form,
        "url": "/transfer",
        "header": header,
        "href": "/store"
    }
    return HttpResponse(template.render(context, request))

def generate_acceptance_act_xls(tmplt_path, res_path, data):
    wb = openpyxl.load_workbook(tmplt_path)
    pg = wb['Sheet1']
    pg['A1'] = f"Акт приема сырья № {data['num']} от {data['date']}"
    pg['B2'] = data["receiver"]
    pg['B3'] = data["sender"]
    pg['B4'] = data["car"]
    pg['B5'] = data["driver"]
    pg['B6'] = data["feature"] if not data["feature2"] else f'{data["feature"]}, {data["feature2"]}'
    pg['B7'] = data["waybill_num"]
    pg['B8'] = data["waybill_weight"]
    total = data["kip_count"]
    for i in range(len(data["weights"])):
        if data["feature2"] and total <= 0:
            data["penal_count"] = data["penal_count2"]
            total = data["kip_count2"]
            data["feature"] = data["feature2"]
        count = data["penal_count"] if total - data["penal_count"] > 0 else total
        pg.append([count, data["feature"], " ", data["weights"][i], " ", data["platform"]])
        total -= data["penal_count"]
    pg.append([" ", " ", " ", " "])
    pg.append([" ", " ", " ", " "])
    pg.append([data["kip_count"], "ИТОГО", " ", data["in_weight"] - data["out_weight"], " ", data["platform"]])
    pg.append(["Вес автомобиля на въезде:", data["in_weight"], " ", " "])
    pg.append(["Вес автомобиля на выезде:", data["out_weight"], "ИТОГО", data["in_weight"] - data["out_weight"], " "])
    pg.append([" ", "ПРИНЯЛ", " ", " ", " ", data["receiving_worker"]])
    wb.save(res_path)


def generate_acceptance_act(request, id, fn):
    if not request.user.is_authenticated:
        return redirect("/login")
    tmplt_path = os.path.join(os.path.dirname(__file__), 'static', 'files', 'acceptance_act_template.xlsx')
    res_path = os.path.join(os.path.dirname(__file__), 'static', 'files', 'acceptance_act.xlsx')
    act = get_object_or_404(AcceptanceAct, pk=id)
    data = {}
    data["num"] = act.act_num
    act_date = act.date + timedelta(hours=3)
    data["date"] = act_date.strftime("%d.%m.%Y")
    data["receiver"] = f'{act.receiver.type}"{act.receiver.name}"'
    data["sender"] = f'{act.sender.type}"{act.sender.name}"'
    data["car"] = f"{act.car.mark} {act.car.reg_num}"
    data["driver"] = str(act.driver)
    data["feature"] = str(act.raw_material)
    data["feature2"] = str(act.raw_material2)
    data["waybill_num"] = act.waybill_num
    data["waybill_weight"] = act.waybill_weight
    data["weights"] = act.weight_list
    data["penal_count"] = act.penal_count
    data["kip_count"] = act.kip_count
    data["penal_count2"] = act.penal_count2
    data["kip_count2"] = act.kip_count2
    data["platform"] = f'{act.platform.type}"{act.platform.name}"'
    data["in_weight"] = act.in_weight
    data["out_weight"] = act.out_weight
    data["receiving_worker"] = str(act.receiving_worker)
    generate_acceptance_act_xls(tmplt_path, res_path, data)
    with open(res_path, "rb") as file:
        response = HttpResponse(content=file)
        response['Content-Type'] = 'application/xlsx'
    return response

def generate_brigade_order_docx(tmplt_path, res_path, data):
    doc = DocxTemplate(tmplt_path)
    doc.render(data)
    doc.save(res_path)

def generate_brigade_order(request, id, fn):
    if not request.user.is_authenticated:
        return redirect("/login")
    tmplt_path = os.path.join(os.path.dirname(__file__), 'static', 'files', 'brigade_order_tmplt.docx')
    res_path = os.path.join(os.path.dirname(__file__), 'static', 'files', 'brigade_order.docx')
    order = Waste.objects.get(pk=id)
    data = {}
    data["black_metal"] = order.black_metal
    data["color_metal"] = order.color_metal
    data["off_color_bottle"] = order.off_color_bottle
    data["pss"] = order.pss
    data["paper"] = order.paper
    data["other"] = order.other
    black_metal_start = 0
    color_metal_start = 0
    off_color_bottle_start = 0
    pss_start = 0
    paper_start = 0
    other_start = 0
    for older_order in Waste.objects.filter(date__lt=order.date):
        if order.date.m == older_order.date.m:
            black_metal_start += older_order.black_metal
            color_metal_start += older_order.color_metal
            off_color_bottle_start += older_order.off_color_bottle
            pss_start += older_order.pss
            paper_start += older_order.paper
            other_start += older_order.other
    data["black_metal_start"] = black_metal_start
    data["color_metal_start"] = color_metal_start
    data["off_color_bottle_start"] = off_color_bottle_start
    data["pss_start"] = pss_start
    data["paper_start"] = paper_start
    data["other_start"] = other_start
    data["brigade_num"] = order.brigade_num
    data["total_start"] = black_metal_start + color_metal_start + off_color_bottle_start + pss_start + paper_start + other_start
    data["total"] = order.black_metal + order.color_metal + order.off_color_bottle + order.pss + order.paper + order.otherghjb
    generate_brigade_order_docx(tmplt_path, res_path, data)
    with open(res_path, "rb") as file:
        response = HttpResponse(content=file)
        response['Content-Type'] = 'application/docx'
    return response

def brigadeorder(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    doc_data = []
    template = loader.get_template("management/brigadeorder.html")
    for waste in Waste.objects.all():
        date = waste.date + timedelta(hours=3)
        waste.date = date.strftime('%d.%m.%Y')
        waste.generate_link = f"brigadeorder/{waste.id}/generate/order{date.strftime('%d.%m.%Y')}.docx"
        doc_data.append(waste)
    doc_data.sort(key=lambda x: x.date, reverse=True)
    context = {
        "doc_data": doc_data,
        "user": Users.objects.filter(username=request.user.username).get()
    }
    return HttpResponse(template.render(context, request))

def organization(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/form.html")
    header = "Добавить организацию"
    if request.method == "POST":
        form = OrganizationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("handbook")
        else:
            header = "Ошибка при сохранении"
    bank_form = OrganizationForm(initial={'is_default': False})
    context = {
        "form": bank_form,
        "url": "organization",
        "header": header,
        "href": "handbook"
    }
    return HttpResponse(template.render(context, request))

def edit_organization(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    entity = get_object_or_404(Organization, pk=id)
    template = loader.get_template("management/form.html")
    header = "Изменить организацию"
    if request.method == 'POST':
        form = OrganizationForm(request.POST, instance=entity)
        form.save()
        return redirect("handbook")
    form = OrganizationForm(instance=entity)
    context = {
        "form": form,
        "url": "/organization/" + str(entity.id),
        "header": header,
        "href": "/handbook"
    }
    return HttpResponse(template.render(context, request))

def rawmaterial(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/form.html")
    header = "Добавить вид сырья"
    if request.method == "POST":
        form = RawMaterialForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("handbook")
        else:
            header = "Ошибка при сохранении"
    bank_form = RawMaterialForm()
    context = {
        "form": bank_form,
        "url": "rawmaterial",
        "header": header,
        "href": "handbook"
    }
    return HttpResponse(template.render(context, request))

def edit_rawmaterial(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    entity = get_object_or_404(RawMaterial, pk=id)
    template = loader.get_template("management/form.html")
    header = "Изменить вид сырья"
    if request.method == 'POST':
        form = RawMaterialForm(request.POST, instance=entity)
        form.save()
        return redirect("handbook")
    form = RawMaterialForm(instance=entity)
    context = {
        "form": form,
        "url": "/rawmaterial/" + str(entity.id),
        "header": header,
        "href": "/handbook"
    }
    return HttpResponse(template.render(context, request))

def car(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/form.html")
    header = "Добавить автомобиль"
    if request.method == "POST":
        form = CarForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("handbook")
        else:
            header = "Ошибка при сохранении"
    bank_form = CarForm()
    context = {
        "form": bank_form,
        "url": "car",
        "header": header,
        "href": "handbook"
    }
    return HttpResponse(template.render(context, request))

def edit_car(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    entity = get_object_or_404(Car, pk=id)
    template = loader.get_template("management/form.html")
    header = "Изменить автомобиль"
    if request.method == 'POST':
        form = CarForm(request.POST, instance=entity)
        form.save()
        return redirect("handbook")
    form = CarForm(instance=entity)
    context = {
        "form": form,
        "url": "/car/" + str(entity.id),
        "header": header,
        "href": "/handbook"
    }
    return HttpResponse(template.render(context, request))

def driver(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/form.html")
    header = "Добавить водителя"
    if request.method == "POST":
        form = DriverForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("handbook")
        else:
            header = "Ошибка при сохранении"
    bank_form = DriverForm()
    context = {
        "form": bank_form,
        "url": "driver",
        "header": header,
        "href": "handbook"
    }
    return HttpResponse(template.render(context, request))

def edit_driver(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    entity = get_object_or_404(Driver, pk=id)
    template = loader.get_template("management/form.html")
    header = "Изменить водителя"
    if request.method == 'POST':
        form = DriverForm(request.POST, instance=entity)
        form.save()
        return redirect("handbook")
    form = DriverForm(instance=entity)
    context = {
        "form": form,
        "url": "/driver/" + str(entity.id),
        "header": header,
        "href": "/handbook"
    }
    return HttpResponse(template.render(context, request))

def worker(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/form.html")
    header = "Добавить сотрудника"
    if request.method == "POST":
        form = WorkerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("handbook")
        else:
            header = "Ошибка при сохранении"
    bank_form = WorkerForm()
    context = {
        "form": bank_form,
        "url": "worker",
        "header": header,
        "href": "handbook"
    }
    return HttpResponse(template.render(context, request))

def edit_worker(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    entity = get_object_or_404(Worker, pk=id)
    template = loader.get_template("management/form.html")
    header = "Изменить сотрудника"
    if request.method == 'POST':
        form = WorkerForm(request.POST, instance=entity)
        form.save()
        return redirect("handbook")
    form = WorkerForm(instance=entity)
    context = {
        "form": form,
        "url": "/worker/" + str(entity.id),
        "header": header,
        "href": "/handbook"
    }
    return HttpResponse(template.render(context, request))

def productfeature(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/form.html")
    header = "Добавить характеристику продукции"
    if request.method == "POST":
        form = ProductFeatureForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("handbook")
        else:
            header = "Ошибка при сохранении"
    bank_form = ProductFeatureForm()
    context = {
        "form": bank_form,
        "url": "productfeature",
        "header": header,
        "href": "handbook"
    }
    return HttpResponse(template.render(context, request))

def edit_productfeature(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    entity = get_object_or_404(ProductFeature, pk=id)
    template = loader.get_template("management/form.html")
    header = "Изменить характеристику продукции"
    if request.method == 'POST':
        form = ProductFeatureForm(request.POST, instance=entity)
        form.save()
        return redirect("handbook")
    form = ProductFeatureForm(instance=entity)
    context = {
        "form": form,
        "url": "/productfeature/" + str(entity.id),
        "header": header,
        "href": "/handbook"
    }
    return HttpResponse(template.render(context, request))

def productnom(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/form.html")
    header = "Добавить номенклатуру продукции"
    if request.method == "POST":
        form = ProductNomForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("handbook")
        else:
            header = "Ошибка при сохранении"
    bank_form = ProductNomForm()
    context = {
        "form": bank_form,
        "url": "productnom",
        "header": header,
        "href": "handbook"
    }
    return HttpResponse(template.render(context, request))

def edit_productnom(request, id):
    if not request.user.is_authenticated:
        return redirect("/login")
    entity = get_object_or_404(ProductNom, pk=id)
    template = loader.get_template("management/form.html")
    header = "Изменить номенклатуру продукции"
    if request.method == 'POST':
        form = ProductNomForm(request.POST, instance=entity)
        form.save()
        return redirect("handbook")
    form = ProductNomForm(instance=entity)
    context = {
        "form": form,
        "url": "/productnom/" + str(entity.id),
        "header": header,
        "href": "/handbook"
    }
    return HttpResponse(template.render(context, request))

def journal(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/journal.html")
    journal_data = []
    for act in AcceptanceAct.objects.filter(status__in=["done", "archive"]):
        remain_count = 0
        remain_weight = 0
        act_date = act.date + timedelta(hours=3)
        data = {
            "date": act_date.strftime("%d.%m.%Y"),
            "act": f"{act.act_num} {Organization.objects.get(pk=act.sender_id).name}",
            "nom": RawMaterial.objects.get(pk=act.raw_material_id).feature,
            "rest_count": remain_count,
            "rest_weight": remain_weight,
        }
        remain_weight += round(sum(act.weight_list))
        remain_count += int(act.penal_count * len(act.weight_list))
        data["arrival_count"] = remain_count
        data["arrival_weight"] = remain_weight
        data["flow_count"] = 0
        data["flow_weight"] = 0
        data["final_rest_count"] = remain_count
        data["final_rest_weight"] = remain_weight
        journal_data.append(data)
        for transfer in TransferToProd.objects.filter(act_id=act.id).order_by("date"):
            transfer_date = transfer.date + timedelta(hours=3)
            data = {
                "date": transfer_date.strftime("%d.%m.%Y"),
                "act": f"{act.act_num} {Organization.objects.get(pk=act.sender_id).name}",
                "nom": RawMaterial.objects.get(pk=act.raw_material_id).feature,
                "rest_count": remain_count,
                "rest_weight": remain_weight,
            }
            remain_weight -= round((sum(act.weight_list) / (act.kip_count if not act.kip_count2 else act.kip_count + act.kip_count2)) * transfer.transfer_count)
            remain_count -= int(transfer.transfer_count)
            data["arrival_count"] = 0
            data["arrival_weight"] = 0
            data["flow_count"] = transfer.transfer_count
            data["flow_weight"] = round((sum(act.weight_list) / (act.kip_count if not act.kip_count2 else act.kip_count + act.kip_count2)) * transfer.transfer_count)
            data["final_rest_count"] = remain_count
            data["final_rest_weight"] = round(remain_weight, 2)
            journal_data.append(data)
    journal_data.sort(key=lambda x: x["date"])
    context = {
        "user": Users.objects.filter(username=request.user.username).get(),
        "cur_date": timezone.now().strftime("%d.%m.%Y"),
        "journal_data": journal_data,
        "transfer_link": f"/transfer/transfer_to_prod{timezone.now().day}.{timezone.now().month}.{timezone.now().year}.xlsx"
    }
    return HttpResponse(template.render(context, request))

def waste(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/form.html")
    header = "Добавить возвратные отходы"
    if request.method == "POST":
        form = WasteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("store")
        else:
            header = "Ошибка при сохранении"
    form = WasteForm()
    context = {
        "form": form,
        "url": "waste",
        "header": header,
        "href": "store"
    }
    return HttpResponse(template.render(context, request))

def productspecification(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/form.html")
    header = "Добавить спецификацию продукции"
    if request.method == "POST":
        form = ProductSpecificationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("/production/documents")
        else:
            header = "Ошибка при сохранении"
    form = ProductSpecificationForm()
    context = {
        "form": form,
        "url": "productspecification",
        "header": header,
        "href": "/production/documents"
    }
    return HttpResponse(template.render(context, request))

def materials(request):
    if not request.user.is_authenticated:
        return redirect("/login")
    template = loader.get_template("management/materials.html")
    docs = []
    for act in AcceptanceAct.objects.filter(status__in=["done", "archive"]):
        act_data = {}
        act_data["act_num"] = act.act_num
        act_data["sender"] = act.sender
        act_date = act.date + timedelta(hours=3)
        act_data["date"] = act_date.strftime("%d.%m.%Y")
        if act.raw_material2:
            act_data["material"] = ", ".join([str(act.raw_material), str(act.raw_material2)])
        else:
            act_data["material"] = str(act.raw_material)
        act_data["average_weight"] = sum(act.weight_list) / (act.kip_count if not act.kip_count2 else act.kip_count + act.kip_count2)
        act_data["weight"] = sum(act.weight_list)
        act_data["count"] = act.kip_count if not act.kip_count2 else act.kip_count + act.kip_count2
        if TransferToProd.objects.filter(act_id=act.id).exists():
            for transfer in TransferToProd.objects.filter(act_id=act.id):
                act_data["weight"] -= (sum(act.weight_list) / (act.kip_count if not act.kip_count2 else act.kip_count + act.kip_count2)) * transfer.transfer_count
                act_data["count"] -= transfer.transfer_count
        else:
            pass
        if act_data["weight"] > 10:
            docs.append(act_data)
    total = sum(map(lambda x: int(x["weight"]), docs))
    context = {
        "user": Users.objects.filter(username=request.user.username).get(),
        "docs": docs,
        "total": total
    }
    return HttpResponse(template.render(context, request))


def login_view(request):
    error_status = ""
    if request.method == 'POST':
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("/")
        else:
            error_status = "Неверный логин или пароль"
    template = loader.get_template("management/login.html")
    context = {
        "error_status": error_status
    }
    return HttpResponse(template.render(context, request))

def logout_view(request):
    logout(request)
    return redirect("/login")

